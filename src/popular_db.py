import mysql.connector
from openpyxl import load_workbook
from datetime import datetime


# Função para tratar os valores monetários
def tratar_valor(valor_str):
    # Remover o símbolo da moeda e os pontos de milhares
    valor_str = valor_str.replace("R$", "").replace(".", "")
    # Substituir a vírgula por ponto para separar os centavos
    valor_str = valor_str.replace(",", ".")
    # Converter para float
    return float(valor_str)

# Conectando ao banco de dados
conexao = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="resultado_mega"
)

# Criando um cursor
cursor = conexao.cursor()

# Abrindo o arquivo de log para ler o nome do arquivo a ser consultado
with open('log.txt', 'r') as log_file:
    lines = log_file.readlines()
    if lines:
        ultimo_arquivo = lines[-1].strip()  # Obtendo o último nome de arquivo no log
    else:
        print("Log está vazio.")
        exit()

# Carregando o arquivo Excel
workbook = load_workbook(filename=f'resultados/{ultimo_arquivo}')
sheet = workbook.active

# Inicializando o número de concurso
num_concurso = 1

# Iterando pelas linhas do arquivo Excel e inserindo os dados na tabela
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Convertendo a data do sorteio para o formato de objeto de data do Python
    data_sorteio = datetime.strptime(row[1], '%d/%m/%Y').date()
    
    # Tratando os valores monetários
    rateio_6_acertos = tratar_valor(row[10])
    rateio_5_acertos = tratar_valor(row[12])
    rateio_4_acertos = tratar_valor(row[14])
    acumulado_6_acertos = tratar_valor(row[15])
    arrecadacao_total = tratar_valor(row[16])
    estimativa_premio = tratar_valor(row[17])
    acumulado_virada = tratar_valor(row[18])
    
    comando_sql = """
    INSERT INTO geral (
        Concurso, Data_do_Sorteio, Bola1, Bola2, Bola3, Bola4, Bola5, Bola6,
        Ganhadores_6_acertos, Cidade_UF, Rateio_6_acertos, Ganhadores_5_acertos,
        Rateio_5_acertos, Ganhadores_4_acertos, Rateio_4_acertos,
        Acumulado_6_acertos, Arrecadacao_Total, Estimativa_premio,
        Acumulado_Sorteio_Especial_Mega_da_Virada, Observacao
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    dados = (
        num_concurso, data_sorteio, row[2], row[3], row[4], row[5], row[6],
        row[7], row[8], row[9], rateio_6_acertos, row[11], rateio_5_acertos, row[13], 
        rateio_4_acertos, acumulado_6_acertos, arrecadacao_total, estimativa_premio, 
        acumulado_virada, row[19]
    )
    cursor.execute(comando_sql, dados)
    num_concurso += 1  # Incrementando o número de concurso

# Commitando as alterações
conexao.commit()

# Fechando o cursor e a conexão
cursor.close()
conexao.close()

print("Dados inseridos com sucesso.")
